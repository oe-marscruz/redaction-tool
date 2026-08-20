"""Launch the Redaction Tool desktop app.

Usage:
    python run.py
    python -m redaction_tool
    python run.py --selftest <output-file>   # headless validation, writes OK/FAIL
"""

import sys


def _selftest(outfile: str) -> int:
    """End-to-end validation without a GUI.

    Builds small sample PDF/DOCX/XLSX files containing known PII, redacts
    them, re-extracts the text, and verifies the PII is gone.
    """
    import tempfile
    from pathlib import Path

    from redaction_tool import redactor

    pii_strings = ["Brianna Ynostroza", "jane.doe@example.edu", "123-45-6789"]
    report: list[str] = []

    with tempfile.TemporaryDirectory() as tmp:
        tmpdir = Path(tmp)
        samples: list[Path] = []

        # --- DOCX sample ---
        try:
            import docx
            d = docx.Document()
            d.add_paragraph(
                "Student Brianna Ynostroza (SSN 123-45-6789) may be reached "
                "at jane.doe@example.edu or (303) 555-0142. DOB 05/14/2002."
            )
            p = tmpdir / "sample.docx"
            d.save(str(p))
            samples.append(p)
            report.append("DOCX sample created")
        except Exception as exc:  # noqa: BLE001
            report.append(f"DOCX setup failed: {exc}")

        # --- XLSX sample ---
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = "Name"
            ws["B1"] = "Brianna Ynostroza"
            ws["A2"] = "Email"
            ws["B2"] = "jane.doe@example.edu"
            ws["A3"] = "SSN"
            ws["B3"] = "123-45-6789"
            p = tmpdir / "sample.xlsx"
            wb.save(str(p))
            samples.append(p)
            report.append("XLSX sample created")
        except Exception as exc:  # noqa: BLE001
            report.append(f"XLSX setup failed: {exc}")

        # --- PDF sample ---
        try:
            try:
                import pymupdf as fitz
            except ImportError:
                import fitz
            doc = fitz.open()
            page = doc.new_page()
            page.insert_text(
                (72, 72),
                "Student Brianna Ynostroza\nSSN 123-45-6789\n"
                "Email jane.doe@example.edu\nPhone (303) 555-0142",
            )
            p = tmpdir / "sample.pdf"
            doc.save(str(p))
            doc.close()
            samples.append(p)
            report.append("PDF sample created")
        except Exception as exc:  # noqa: BLE001
            report.append(f"PDF setup failed: {exc}")

        # --- redact and verify ---
        opts = redactor.ScanOptions()  # all categories
        for sample in samples:
            out = sample.with_name(sample.stem + "_REDACTED" + sample.suffix)
            result = redactor.redact_file(sample, opts, out_path=out)
            if result.error:
                report.append(f"FAIL {sample.name}: {result.error}")
                continue
            redacted_text = redactor.extract_text(out)
            leaks = [s for s in pii_strings if s.lower() in redacted_text.lower()]
            if leaks:
                report.append(f"FAIL {sample.name}: leaked {leaks}")
            else:
                report.append(f"PASS {sample.name}: {result.redaction_count} redactions")

        # --- image rendering check ---
        pdf_out = tmpdir / "sample_REDACTED.pdf"
        if pdf_out.exists():
            try:
                imgs = redactor._pdf_to_images(pdf_out, tmpdir / "imgs", "PNG")
                report.append(f"PASS image render: {len(imgs)} page(s)")
            except Exception as exc:  # noqa: BLE001
                report.append(f"FAIL image render: {exc}")

    ok = all(line.startswith(("PASS", "DOCX sample", "XLSX sample", "PDF sample"))
             for line in report) and any(l.startswith("PASS") for l in report)
    with open(outfile, "w", encoding="utf-8") as f:
        f.write(("OK" if ok else "FAIL") + "\n\n" + "\n".join(report))
    return 0 if ok else 1


def main() -> int:
    if len(sys.argv) >= 3 and sys.argv[1] == "--selftest":
        return _selftest(sys.argv[2])
    from redaction_tool.gui import main as gui_main
    gui_main()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
